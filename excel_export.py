"""Admin uchun Excel hisobotlarini yaratish."""
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


def _style_header(ws, ncols):
    header_fill = PatternFill(start_color="FF9F1C", end_color="FF9F1C", fill_type="solid")
    header_font = Font(bold=True, color="1A1A1A")
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    # Ustun kengligini avtomatik moslash
    for col in ws.columns:
        max_len = 0
        letter = col[0].column_letter
        for cell in col:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[letter].width = min(max_len + 3, 40)


def make_users_excel(rows) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Foydalanuvchilar"
    headers = ["ID", "Ismi", "Telegram ID", "Furalar", "Reyslar", "Daromad ($)", "Reys xarajati", "Ta'mirlash", "Sof foyda ($)"]
    ws.append(headers)
    for r in rows:
        uid, name, tg, trucks, trips, income, trip_exp, repair = r
        net = income - trip_exp - repair
        ws.append([uid, name, tg, trucks, trips, round(income, 2), round(trip_exp, 2), round(repair, 2), round(net, 2)])
    _style_header(ws, len(headers))
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def make_trucks_excel(rows) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Moshinalar"
    headers = ["ID", "Nomi", "Egasi", "Telegram ID", "Qo'shilgan sana", "Ta'mirlash ($)"]
    ws.append(headers)
    for r in rows:
        tid, name, owner, tg, created, repair = r
        ws.append([tid, name, owner, tg, created, round(repair, 2)])
    _style_header(ws, len(headers))
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def make_trips_excel(rows) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Reyslar"
    headers = ["ID", "Fura", "Egasi", "Holati", "Boshlangan", "Tugagan", "Daromad ($)", "Xarajat ($)", "Foyda ($)"]
    ws.append(headers)
    for r in rows:
        tid, truck, owner, status, created, finished, income, expense = r
        status_uz = "Faol" if status == "active" else "Tugagan"
        profit = income - expense
        ws.append([tid, truck, owner, status_uz, created, finished or "—", round(income, 2), round(expense, 2), round(profit, 2)])
    _style_header(ws, len(headers))
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
